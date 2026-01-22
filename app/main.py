import sys
import os
import traceback
from flask import Flask, request, jsonify, send_from_directory, send_file

# 1. Initialize Flask App IMMEDIATELY
app = Flask(__name__, static_folder='static')
app.config["JSON_SORT_KEYS"] = False

# 2. Define Global Error State
STARTUP_ERROR = None
cors_enabled = False

# 3. Safe Imports
try:
    # Flask CORS
    try:
        from flask_cors import CORS
        CORS(app)
        cors_enabled = True
    except ImportError:
        pass
    except Exception:
        pass

    # Standard Libs
    from datetime import date, datetime
    import io
    import csv
    import json
    import tempfile
    
    # Third Party
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape, A3
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    # SQLAlchemy
    from sqlalchemy.orm import Session
    from sqlalchemy.exc import OperationalError
    from sqlalchemy import select, distinct, func

    # Local Imports
    from .database import Base, engine, get_db
    from . import models, schemas, crud, auth, database, migrations
    from .seeds import NIGERIA_STATES_LGAS, seed_special_admin

    # Run migrations on startup
    try:
        migrations.run_migrations()
        # Ensure special admin exists
        with next(get_db()) as db:
            seed_special_admin(db)
    except Exception as e:
        print(f"Startup tasks failed: {e}")


except Exception as e:
    STARTUP_ERROR = f"Startup Error: {str(e)}\n{traceback.format_exc()}"
    print(STARTUP_ERROR)
    
    Base = None
    engine = None
    get_db = lambda: (yield None)
    models = None
    schemas = None
    crud = None
    auth = None
    database = None
    NIGERIA_STATES_LGAS = {}
    class Font: pass
    class Alignment: pass
    class PatternFill: pass

@app.route("/ping")
def ping():
    if STARTUP_ERROR:
        return jsonify({"status": "error", "message": STARTUP_ERROR}), 500
    return "pong"

# --- AUTH ---
@app.route("/login.html")
def login_page():
    # Robust static file serving
    try:
        return send_from_directory(app.static_folder, "login.html")
    except Exception as e:
        return f"Static file error: {e}. Static folder: {app.static_folder}, CWD: {os.getcwd()}", 404

@app.route("/success.html")
def success_page():
    return send_from_directory(app.static_folder, "success.html")

@app.post("/login")
def login():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    try:
        data = request.get_json(force=True)
        username = data.get("username")
        password = data.get("password")
        
        if not username or not password:
            return jsonify({"detail": "Username and password required"}), 400

        def attempt_login():
            with next(get_db()) as db:
                if not db: raise Exception("Database connection failed")
                user = db.query(models.User).filter(models.User.username == username).first()
                if user:
                    # Robust password verification with auto-fix for legacy/bad hashes
                    verification_success = False
                    try:
                        if auth.verify_password(password, user.password_hash):
                            verification_success = True
                    except ValueError:
                        # Handle "hash could not be identified" (e.g. legacy bcrypt vs pbkdf2)
                        # If it's the admin user, we auto-heal the password
                        if user.username == "admin":
                            print(f"WARNING: fixing invalid password hash for admin")
                            user.password_hash = auth.get_password_hash("admin")
                            db.commit()
                            if auth.verify_password(password, user.password_hash):
                                verification_success = True
                    
                    if verification_success:
                        token = auth.create_access_token(data={
                            "sub": user.username, 
                            "role": user.role, 
                            "id": user.id,
                            "organization_id": user.organization_id
                        })
                        return jsonify({
                            "access_token": token, 
                            "token_type": "bearer", 
                            "role": user.role, 
                            "username": user.username, 
                            "id": user.id,
                            "organization_id": user.organization_id
                        })
                
                staff = crud.get_staff_by_nis(db, username)
                if staff:
                    # Check global login permission for staff role
                    if staff.role == "staff" and not getattr(staff, "allow_login", 1):
                        return jsonify({"detail": "Login is currently disabled for staff users."}), 403

                    # Check login limit (skip for admins)
                    if staff.role not in ("office_admin", "super_admin", "main_admin") and staff.login_count >= 10:
                        return jsonify({"detail": "Login limit exceeded. Please contact Super Admin to reset."}), 403

                    verification_success = False
                    if staff.password_hash:
                         try:
                             if auth.verify_password(password, staff.password_hash):
                                 verification_success = True
                         except ValueError:
                             pass
                    elif password == staff.nis_no:
                         verification_success = True

                    if verification_success:
                        # Increment login count
                        staff.login_count += 1
                        db.commit()
                        
                        token = auth.create_access_token(data={
                            "sub": staff.nis_no, 
                            "role": staff.role, 
                            "id": staff.id,
                            "organization_id": staff.organization_id
                        })
                        return jsonify({
                            "access_token": token, 
                            "token_type": "bearer", 
                            "role": staff.role, 
                            "username": staff.nis_no, 
                            "id": staff.id,
                            "organization_id": staff.organization_id
                        })
                
                return jsonify({"detail": "Invalid credentials"}), 401

        try:
            return attempt_login()
        except Exception as e:
            msg = str(e).lower()
            if "no such table" in msg or ("relation" in msg and "does not exist" in msg):
                if engine:
                    Base.metadata.create_all(bind=engine)
                    with next(get_db()) as temp_db:
                         from .seeds import seed_default_admin, seed_special_admin
                         seed_default_admin(temp_db)
                         seed_special_admin(temp_db)
                    return attempt_login()
            raise e

    except Exception as e:
        import traceback
        traceback.print_exc()
        # Return the actual error in 'detail' so it's visible in the UI
        return jsonify({
            "detail": f"Server Error: {str(e)}",
            "error": str(e),
            "trace": traceback.format_exc()
        }), 500

@app.get("/me")
def get_current_user_info():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({"detail": "Not authenticated"}), 401
    
    token = auth_header.split(" ")[1]
    payload = auth.decode_access_token(token)
    if not payload:
        return jsonify({"detail": "Invalid token"}), 401
    
    return jsonify(payload)

# --- ORGANIZATION MANAGEMENT (Special Admin) ---

@app.post("/organizations")
def create_organization_endpoint():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["special_admin"])
    if err: return err, code
    
    data = request.get_json()
    name = data.get("name")
    code_val = data.get("code")
    
    if not name or not code_val:
        return jsonify({"detail": "Name and Code are required"}), 400
        
    with next(get_db()) as db:
        try:
            # Check if code exists
            existing = db.query(models.Organization).filter(models.Organization.code == code_val).first()
            if existing:
                return jsonify({"detail": "Organization code already exists"}), 400
                
            org = crud.create_organization(db, name, code_val)
            return jsonify({"id": org.id, "name": org.name, "code": org.code})
        except Exception as e:
            return jsonify({"detail": str(e)}), 400

@app.get("/organizations")
def list_organizations_endpoint():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["special_admin"])
    if err: return err, code
    
    with next(get_db()) as db:
        orgs = crud.list_organizations(db)
        return jsonify([{"id": o.id, "name": o.name, "code": o.code} for o in orgs])

@app.post("/organizations/<int:org_id>/admin")
def create_organization_admin(org_id):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["special_admin"])
    if err: return err, code
    
    data = request.get_json()
    username = data.get("username")
    password = data.get("password")
    
    if not username or not password:
        return jsonify({"detail": "Username and password required"}), 400
        
    with next(get_db()) as db:
        org = crud.get_organization(db, org_id)
        if not org:
            return jsonify({"detail": "Organization not found"}), 404
            
        # Check if user exists
        existing = db.query(models.User).filter(models.User.username == username).first()
        if existing:
            return jsonify({"detail": "Username already taken"}), 400
            
        pwd_hash = auth.get_password_hash(password)
        new_admin = models.User(
            username=username,
            password_hash=pwd_hash,
            role="super_admin", # Organization super admin
            organization_id=org_id
        )
        db.add(new_admin)
        db.commit()
        
        return jsonify({"detail": f"Admin created for {org.name}", "username": username})

@app.get("/dashboard/stats")
def dashboard_stats():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user = get_current_user()
    if not user: return jsonify({"detail": "Not authenticated"}), 401
    
    organization_id = user.get("organization_id")
    
    with next(get_db()) as db:
        if user.get("role") == "office_admin":
            staff_user = crud.get_staff(db, user.get("id"))
            office_name = staff_user.office if staff_user else None
            if not office_name:
                return jsonify({"total_staff": 0, "office_name": None, "rank_counts": {}})

            total_staff = db.scalar(
                select(func.count(models.Staff.id)).where(
                    models.Staff.exit_date.is_(None),
                    models.Staff.office == office_name,
                    models.Staff.organization_id == organization_id
                )
            )

            rank_rows = db.execute(
                select(models.Staff.rank, func.count(models.Staff.id))
                .where(
                    models.Staff.exit_date.is_(None),
                    models.Staff.office == office_name,
                    models.Staff.organization_id == organization_id
                )
                .group_by(models.Staff.rank)
            ).all()
            rank_counts = {}
            for rank, count in rank_rows:
                key = rank or ""
                rank_counts[key] = rank_counts.get(key, 0) + count

            return jsonify({
                "total_staff": total_staff,
                "office_name": office_name,
                "rank_counts": rank_counts,
            })

        stats = crud.get_dashboard_stats(db, organization_id=organization_id)
        stats["office_name"] = None
        return jsonify(stats)

@app.get("/admin/exit-requests")
def list_exit_requests():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "main_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    with next(get_db()) as db:
        stmt = select(models.Staff).where(models.Staff.out_request_status == "Pending")
        if organization_id is not None:
            stmt = stmt.where(models.Staff.organization_id == organization_id)
        stmt = stmt.order_by(models.Staff.out_request_date.asc())
        items = db.scalars(stmt).all()
        return jsonify([schemas.to_dict_staff(item) for item in items])

# --- END AUTH ---

def parse_date_value(value):
    if value is None: return None
    if isinstance(value, date) and not isinstance(value, datetime): return value
    if isinstance(value, datetime): return value.date()
    if isinstance(value, str):
        s = value.strip()
        if not s: return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try: return datetime.strptime(s, fmt).date()
            except ValueError: continue
        try: return date.fromisoformat(s.split("T", 1)[0])
        except ValueError: return None
    return None

def get_current_user():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '): return None
    token = auth_header.split(" ")[1]
    return auth.decode_access_token(token)

def require_role(allowed_roles):
    user = get_current_user()
    if not user: return None, jsonify({"detail": "Not authenticated"}), 401
    role = user.get("role")
    # Treat main_admin as equivalent to super_admin for permissions
    effective_role = role
    if role == "main_admin" and "super_admin" in allowed_roles:
        effective_role = "super_admin"
    if effective_role not in allowed_roles: 
        print(f"PERMISSION DENIED: User role '{role}' not in {allowed_roles}")
        return None, jsonify({"detail": f"Permission denied (Role: {role})"}), 403
    return user, None, None

@app.route("/")
def index():
    if STARTUP_ERROR:
        return f"<h1>System Error</h1><pre>{STARTUP_ERROR}</pre>"
    # Try to serve index.html, but fall back to debug info if missing
    try:
        return send_from_directory(app.static_folder, "index.html")
    except Exception as e:
        return f"<h1>VSS API is Running</h1><p>Static file error: {e}</p><p>Static Folder: {app.static_folder}</p><p>CWD: {os.getcwd()}</p><p><a href='/login.html'>Login Page</a></p>"

@app.route("/debug-db")
def debug_db():
    if STARTUP_ERROR: return jsonify({"status": "error", "message": STARTUP_ERROR}), 500
    if not engine:
        return jsonify({"status": "error", "detail": "Database engine not initialized. Check logs."}), 500
        
    try:
        from sqlalchemy import inspect, text
        inspector = inspect(engine)
        tables = inspector.get_table_names()
        
        try:
            with next(get_db()) as db:
                version = db.execute(text("SELECT version()")).scalar()
                if "users" not in tables:
                    Base.metadata.create_all(bind=engine)
                    from .seeds import seed_default_admin, seed_special_admin
                    seed_default_admin(db)
                    seed_special_admin(db)
                    tables = inspector.get_table_names()
        except Exception as query_err:
             return jsonify({
                "status": "error",
                "detail": "Connection successful but query failed",
                "error": str(query_err)
            }), 500

        return jsonify({
            "status": "ok",
            "db_version": version,
            "tables": tables,
            "message": "Connected successfully"
        })
    except Exception as e:
        import traceback
        return jsonify({
            "status": "error",
            "detail": str(e),
            "traceback": traceback.format_exc()
        }), 500

@app.get("/download/template")
def download_template():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Import Template"
    headers = ["NIS/No", "Surname", "Other Names", "Rank", "Gender", "State of Origin", "LGA", "Office", "Phone No", "Qualification", "Home Town", "Next of Kin", "NOK Phone", "Remark", "DOFA", "DOPA", "DOPP", "DOB"]
    ws.append(headers)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="staff_import_template.xlsx")

@app.post("/import/excel")
def import_excel():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["office_admin", "super_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")

    if 'file' not in request.files: return jsonify({"detail": "No file uploaded"}), 400
    file = request.files['file']
    filename = file.filename.lower()
    
    if not filename.endswith(('.xlsx', '.xls')): return jsonify({"detail": "Invalid file type."}), 400
    
    tmp_dir = tempfile.gettempdir()
    tmp_path = os.path.join(tmp_dir, filename)
    file.save(tmp_path)

    try:
        data_rows = []
        wb = None
        if filename.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            col_map = {h: i for i, h in enumerate(headers) if h}
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for h, i in col_map.items(): row_dict[h] = row[i]
                data_rows.append(row_dict)
                            
        success_count = 0
        errors = []

        with next(get_db()) as db_session:
            states_cache = {s.name.lower(): s.id for s in db_session.query(models.State).all()}
            for row_idx, row_dict in enumerate(data_rows, start=1):
                try:
                    def get_val(target_key):
                        if target_key in row_dict: return row_dict[target_key]
                        for k, v in row_dict.items():
                            if k.lower() == target_key.lower(): return v
                        if target_key == "NIS/No":
                            for k in row_dict.keys():
                                if k.lower().replace("_", "").replace("/", "") == "nisno" or k.lower() == "nis": return row_dict[k]
                        return None
                    def get_text_val(key):
                        val = get_val(key)
                        if val is None: return None
                        s = str(val).strip()
                        return s if s else None

                    nis = get_text_val("NIS/No")
                    if not nis: continue
                    
                    data = {
                        "nis_no": nis,
                        "surname": get_text_val("Surname"),
                        "other_names": get_text_val("Other Names"),
                        "rank": get_text_val("Rank"),
                        "gender": get_text_val("Gender"),
                        "office": get_text_val("Office"),
                        "phone_no": get_text_val("Phone No") or get_text_val("Phone"),
                        "qualification": get_text_val("Qualification"),
                        "home_town": get_text_val("Home Town"),
                        "next_of_kin": get_text_val("Next of Kin"),
                        "nok_phone": get_text_val("NOK Phone"),
                        "remark": get_text_val("Remark"),
                        "dofa": parse_date_value(get_val("DOFA")),
                        "dopa": parse_date_value(get_val("DOPA")),
                        "dopp": parse_date_value(get_val("DOPP")),
                        "dob": parse_date_value(get_val("DOB")),
                    }
                    
                    if not data["surname"] or not data["rank"]: raise ValueError("Missing Surname or Rank")
                    if not data["gender"]: raise ValueError("Missing Gender")
                        
                    s_name = get_text_val("State of Origin") or get_text_val("State")
                    if s_name and s_name.lower() in states_cache:
                        data["state_id"] = states_cache[s_name.lower()]
                        l_name = get_text_val("LGA")
                        if l_name:
                            lga_obj = db_session.query(models.LGA).filter(models.LGA.state_id == data["state_id"], models.LGA.name == l_name).first()
                            if not lga_obj:
                                lga_obj = models.LGA(name=l_name, state_id=data["state_id"])
                                db_session.add(lga_obj)
                                db_session.flush()
                            data["lga_id"] = lga_obj.id
                                
                    if organization_id is not None:
                        data["organization_id"] = organization_id

                    crud.create_staff(db_session, data)
                    success_count += 1
                except Exception as e:
                    db_session.rollback()
                    errors.append(f"Row {row_idx}: {str(e)}")
            
            if success_count > 0: crud.create_audit_log(db_session, "IMPORT", filename, f"Imported {success_count} records")
            db_session.commit()
        return jsonify({"message": f"Imported {success_count} records", "errors": errors[:10]})

    except Exception as e:
        return jsonify({"detail": f"Failed to process file: {str(e)}"}), 500
    finally:
        if 'wb' in locals() and wb: wb.close()
        if 'tmp_path' in locals() and os.path.exists(tmp_path):
            try: os.remove(tmp_path)
            except Exception: pass

# Init DB on startup if possible
if engine:
    try:
        Base.metadata.create_all(bind=engine)
        def seed_states_lgas(db: Session):
            existing_states = {s.name: s for s in db.query(models.State).all()}
            for state_name in NIGERIA_STATES_LGAS:
                if state_name not in existing_states:
                    s = models.State(name=state_name)
                    db.add(s)
                    existing_states[state_name] = s
            db.flush()
            existing_states = {s.name: s for s in db.query(models.State).all()}
            existing_lgas = set()
            for l in db.query(models.LGA).all(): existing_lgas.add((l.state_id, l.name))
            for state_name, lgas in NIGERIA_STATES_LGAS.items():
                if state_name in existing_states:
                    st = existing_states[state_name]
                    for lga_name in lgas:
                        if (st.id, lga_name) not in existing_lgas:
                            db.add(models.LGA(name=lga_name, state_id=st.id))
                            existing_lgas.add((st.id, lga_name))
            db.commit()

        def seed_super_admin(db: Session):
            admin = db.query(models.User).filter(models.User.username == "admin").first()
            if not admin:
                pwd_hash = auth.get_password_hash("admin")
                admin = models.User(username="admin", password_hash=pwd_hash, role="super_admin")
                db.add(admin)
                db.commit()
            elif admin.role != "super_admin":
                print(f"UPGRADING admin user from {admin.role} to super_admin")
                admin.role = "super_admin"
                db.commit()

        from .database import SessionLocal
        db = SessionLocal()
        try:
            try:
                migrations.run_migrations()
            except Exception as mig_err:
                print(f"MIGRATION ERROR: {mig_err}")

            seed_states_lgas(db)
            seed_super_admin(db)
            from .seeds import seed_special_admin
            seed_special_admin(db)
        finally:
            db.close()
    except Exception as e:
        print(f"DB Init Warning: {e}")

@app.get("/offices")
def list_offices_route():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user = get_current_user()
    # Optional: if not authenticated, return empty or global?
    # Let's assume offices are protected now, or if public, return only public ones?
    # For now, let's try to get org from user if logged in.
    organization_id = user.get("organization_id") if user else None
    
    with next(get_db()) as db:
        items = crud.list_offices_model(db, organization_id=organization_id)
        existing_names = {i.name.strip().lower() for i in items if i and i.name and i.name.strip()}
        
        # Only sync staff offices if user is logged in and belongs to an org (or is special admin)
        if user:
             stmt = select(distinct(models.Staff.office)).where(models.Staff.office.is_not(None), models.Staff.office != "")
             if organization_id:
                 stmt = stmt.where(models.Staff.organization_id == organization_id)
             stmt = stmt.order_by(models.Staff.office)
             
             staff_office_names = list(db.scalars(stmt))
             added = False
             for name in staff_office_names:
                if not name: continue
                clean = str(name).strip()
                if not clean: continue
                key = clean.lower()
                if key in existing_names: continue
                db.add(models.Office(name=clean, organization_id=organization_id))
                existing_names.add(key)
                added = True
             if added:
                db.commit()
                items = crud.list_offices_model(db, organization_id=organization_id)
                
        return jsonify([schemas.to_dict_office(i) for i in items])

@app.post("/offices")
def create_office_route():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "admin"])
    if err: return err, code
    
    data = request.get_json(force=True)
    name = data.get("name")
    if not name: return jsonify({"detail": "Name is required"}), 400
    
    organization_id = user.get("organization_id")
    
    with next(get_db()) as db:
        try:
            obj = crud.create_office(db, name, organization_id=organization_id)
            return jsonify(schemas.to_dict_office(obj)), 201
        except Exception as e:
            return jsonify({"detail": str(e)}), 400

@app.put("/offices/<int:office_id>")
def update_office_route(office_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "admin"])
    if err: return err, code
    
    data = request.get_json(force=True)
    name = data.get("name")
    if not name: return jsonify({"detail": "Name is required"}), 400
    
    with next(get_db()) as db:
        obj = crud.update_office(db, office_id, name)
        if not obj: return jsonify({"detail": "Not found"}), 404
        return jsonify(schemas.to_dict_office(obj))

@app.delete("/offices/<int:office_id>")
def delete_office_route(office_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "admin"])
    if err: return err, code
    
    with next(get_db()) as db:
        if crud.delete_office(db, office_id):
            return jsonify({"detail": "Deleted"}), 200
        return jsonify({"detail": "Not found"}), 404

@app.get("/states")
def get_states():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    with next(get_db()) as db: return jsonify([schemas.to_dict_state(s) for s in crud.list_states(db)])

@app.get("/states/<int:state_id>/lgas")
def get_lgas(state_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    with next(get_db()) as db: return jsonify([schemas.to_dict_lga(l) for l in crud.list_lgas_by_state(db, state_id)])

@app.get("/staff")
def list_staff_endpoint():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    try:
        user = get_current_user()
        if not user: return jsonify({"detail": "Not authenticated"}), 401
        
        q = request.args.get("q")
        state_id = request.args.get("state_id", type=int)
        lga_id = request.args.get("lga_id", type=int)
        
        # Handle multi-select for rank and office
        rank = [r for r in request.args.getlist("rank") if r.strip()]
        if not rank: rank = None
        
        office = [o for o in request.args.getlist("office") if o.strip()]
        if not office: office = None

        completeness = request.args.get("completeness")
        status = request.args.get("status", "active")
        dopp_order = request.args.get("dopp_order")
        exit_from_raw = request.args.get("exit_from")
        exit_to_raw = request.args.get("exit_to")
        exit_from = parse_date_value(exit_from_raw) if exit_from_raw else None
        exit_to = parse_date_value(exit_to_raw) if exit_to_raw else None
        limit = request.args.get("limit", 100, type=int)
        offset = request.args.get("offset", 0, type=int)
        
        organization_id = user.get("organization_id")
        
        with next(get_db()) as db:
            if user["role"] == "office_admin":
                staff_user = crud.get_staff(db, user["id"])
                if not staff_user or not staff_user.office: return jsonify([]), 200
                office = [staff_user.office]
            items = crud.list_staff(
                db,
                q=q,
                state_id=state_id,
                lga_id=lga_id,
                rank=rank,
                office=office,
                completeness=completeness,
                status=status,
                dopp_order=dopp_order,
                limit=limit,
                offset=offset,
                exit_from=exit_from,
                exit_to=exit_to,
                organization_id=organization_id
            )
            return jsonify([schemas.to_dict_staff(item) for item in items])
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"detail": f"Server Error: {str(e)}", "trace": traceback.format_exc()}), 500

@app.post("/staff")
def create_staff():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, error_response, code = require_role(["office_admin", "super_admin"])
    if error_response: return error_response, code

    data = request.get_json(force=True)
    for k in ["nis_no","surname","other_names","rank"]:
        if k not in data or not str(data[k]).strip(): return jsonify({"detail": f"{k} is required"}), 400
    for k in ("dofa", "dopa", "dopp", "dob", "exit_date"):
        if k in data:
            parsed = parse_date_value(data.get(k))
            if data.get(k) not in (None, "") and parsed is None: return jsonify({"detail": f"Invalid date for {k}"}), 400
            data[k] = parsed
    if "gender" not in data or data["gender"] is None: data["gender"] = ""
    
    organization_id = user.get("organization_id")
    if organization_id:
        data["organization_id"] = organization_id
        
    with next(get_db()) as db:
        if user["role"] == "office_admin":
             staff_user = crud.get_staff(db, user["id"])
             if not staff_user or not staff_user.office:
                 return jsonify({"detail": "Admin has no assigned office"}), 403
             data["office"] = staff_user.office
             data["allow_edit_rank"] = 0
             data["allow_edit_dopp"] = 0

        try:
            obj = crud.create_staff(db, data)
            crud.create_audit_log(db, "CREATE", f"Staff: {obj.nis_no}", "Created new staff", organization_id=organization_id)
            return jsonify(schemas.to_dict_staff(obj)), 201
        except ValueError as e: return jsonify({"detail": str(e)}), 400

@app.get("/staff/<int:staff_id>")
def get_staff(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    with next(get_db()) as db:
        obj = crud.get_staff(db, staff_id)
        if not obj: return jsonify({"detail": "Not found"}), 404
        return jsonify(schemas.to_dict_staff(obj))

@app.put("/staff/<int:staff_id>")
def update_staff(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user = get_current_user()
    if not user: return jsonify({"detail": "Not authenticated"}), 401
    
    organization_id = user.get("organization_id")
    
    data = request.get_json(force=True)
    for k in ("dofa", "dopa", "dopp", "dob", "exit_date"):
        if k in data: data[k] = parse_date_value(data.get(k))
    with next(get_db()) as db:
        existing = crud.get_staff(db, staff_id)
        if not existing: return jsonify({"detail": "Not found"}), 404
        
        # Organization check
        if organization_id and existing.organization_id != organization_id:
            return jsonify({"detail": "Permission denied: Different Organization"}), 403
        
        if user["role"] in ["staff", "office_admin"]:
            # Check ownership for staff, or office match for office_admin
            if user["role"] == "staff" and user["id"] != staff_id:
                return jsonify({"detail": "Permission denied"}), 403
            if user["role"] == "office_admin":
                 admin_staff = crud.get_staff(db, user["id"])
                 if not admin_staff or admin_staff.office != existing.office:
                     return jsonify({"detail": "Permission denied"}), 403

            for restricted in ["office", "role", "exit_date", "exit_mode", "out_request_status"]:
                if restricted in data and data[restricted] != getattr(existing, restricted):
                    return jsonify({"detail": f"Permission denied: Cannot change {restricted}"}), 403
            
            # Staff specific restrictions
            if user["role"] == "staff":
                if "rank" in data and data["rank"] != existing.rank and not getattr(existing, "allow_edit_rank", 0):
                    return jsonify({"detail": "Permission denied: Cannot change rank"}), 403
                if "dopp" in data and data["dopp"] != existing.dopp and not getattr(existing, "allow_edit_dopp", 0):
                    return jsonify({"detail": "Permission denied: Cannot change dopp"}), 403
            
            # Create or Update Edit Request instead of direct update
            json_data = {}
            for k, v in data.items():
                if isinstance(v, (date, datetime)):
                    json_data[k] = v.isoformat()
                else:
                    json_data[k] = v
            
            # Check for existing pending request
            stmt = select(models.StaffEditRequest).where(
                models.StaffEditRequest.staff_id == existing.id,
                models.StaffEditRequest.status == "pending"
            )
            existing_req = db.scalar(stmt)

            if existing_req:
                # Merge new changes into existing request
                current_data = json.loads(existing_req.data)
                current_data.update(json_data)
                existing_req.data = json.dumps(current_data)
                # Update timestamp to show latest activity
                existing_req.created_at = func.now()
                
                db.commit()
                crud.create_audit_log(
                    db,
                    "UPDATE_REQUEST_APPEND",
                    f"Staff: {existing.nis_no}",
                    f"Appended to EDIT_REQUEST_ID={existing_req.id}",
                    organization_id=organization_id
                )
                return jsonify({"detail": "Update appended to pending request", "status": "pending_approval"}), 202
            else:
                req = models.StaffEditRequest(
                    staff_id=existing.id,
                    data=json.dumps(json_data),
                    status="pending"
                )
                db.add(req)
                db.commit()
                crud.create_audit_log(
                    db,
                    "UPDATE_REQUEST",
                    f"Staff: {existing.nis_no}",
                    f"EDIT_REQUEST_ID={req.id}",
                    organization_id=organization_id
                )
                return jsonify({"detail": "Update submitted for approval", "status": "pending_approval"}), 202
        

        try:
            obj = crud.update_staff(db, existing, data)
            if obj:
                crud.create_audit_log(db, "UPDATE", f"Staff: {obj.nis_no}", "Updated staff details", organization_id=organization_id)
                return jsonify(schemas.to_dict_staff(obj))
            return jsonify({"detail": "Not found"}), 404
        except ValueError as e:
            return jsonify({"detail": str(e)}), 400

@app.delete("/staff/<int:staff_id>")
def delete_staff(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    with next(get_db()) as db:
        obj = crud.get_staff(db, staff_id)
        if not obj:
            return jsonify({"detail": "Not found"}), 404
            
        if organization_id and obj.organization_id != organization_id:
            return jsonify({"detail": "Permission denied: Different Organization"}), 403
            
        crud.delete_staff(db, obj)
        crud.create_audit_log(db, "DELETE", f"Staff ID: {staff_id}", "Deleted staff record", organization_id=organization_id)
        return jsonify({"detail": "Deleted"})

@app.post("/staff/<int:staff_id>/reset-login")
def reset_login_count(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    with next(get_db()) as db:
        obj = crud.get_staff(db, staff_id)
        if not obj: return jsonify({"detail": "Not found"}), 404
        
        if organization_id and obj.organization_id != organization_id:
            return jsonify({"detail": "Permission denied: Different Organization"}), 403
        
        obj.login_count = 0
        db.add(obj)
        db.commit()
        crud.create_audit_log(db, "RESET_LOGIN", f"Staff: {obj.nis_no}", "Reset login count", organization_id=organization_id)
        return jsonify({"detail": "Login count reset successfully"})

@app.post("/staff/<int:staff_id>/reset-password")
def reset_staff_password(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    with next(get_db()) as db:
        obj = crud.get_staff(db, staff_id)
        if not obj: return jsonify({"detail": "Not found"}), 404
        
        if organization_id and obj.organization_id != organization_id:
            return jsonify({"detail": "Permission denied: Different Organization"}), 403
        
        obj.password_hash = None # Reset to use NIS number
        db.add(obj)
        db.commit()
        crud.create_audit_log(db, "RESET_PASSWORD", f"Staff: {obj.nis_no}", "Reset password to default", organization_id=organization_id)
        return jsonify({"detail": "Password reset successfully"})

@app.put("/staff/<int:staff_id>/role")
def update_staff_role(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    data = request.get_json(force=True)
    new_role = data.get("role")
    if new_role not in ("staff", "office_admin", "super_admin", "main_admin"):
        return jsonify({"detail": "Invalid role"}), 400
    with next(get_db()) as db:
        obj = crud.get_staff(db, staff_id)
        if not obj: return jsonify({"detail": "Not found"}), 404
        
        if organization_id and obj.organization_id != organization_id:
            return jsonify({"detail": "Permission denied: Different Organization"}), 403
            
        obj.role = new_role
        db.add(obj)
        db.commit()
        db.refresh(obj)
        crud.create_audit_log(db, "ROLE_UPDATE", f"Staff: {obj.nis_no}", f"Role set to {new_role}", organization_id=organization_id)
        return jsonify(schemas.to_dict_staff(obj))

@app.post("/staff/<int:staff_id>/move")
def move_staff(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "main_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    data = request.get_json(force=True)
    new_office = data.get("office")
    effective_date_str = data.get("date") # Optional effective date, default today?
    remarks = data.get("remarks", "")

    if not new_office:
        return jsonify({"detail": "New office is required"}), 400
    
    effective_date = date.today()
    if effective_date_str:
        effective_date = parse_date_value(effective_date_str) or date.today()

    with next(get_db()) as db:
        staff = crud.get_staff(db, staff_id)
        if not staff: return jsonify({"detail": "Not found"}), 404
        
        if organization_id and staff.organization_id != organization_id:
            return jsonify({"detail": "Permission denied: Different Organization"}), 403
        
        old_office = staff.office
        if old_office == new_office:
             return jsonify({"detail": "Staff is already in this office"}), 400

        # Create History Record
        history = models.PostingHistory(
            staff_id=staff.id,
            action_type="MOVE",
            from_office=old_office,
            to_office=new_office,
            action_date=effective_date,
            remarks=remarks
        )
        db.add(history)
        
        # Update Staff
        staff.office = new_office
        # Usually a move implies updating DOPP (Date of Present Posting)
        staff.dopp = effective_date
        
        db.commit()
        crud.create_audit_log(db, "MOVE", f"Staff: {staff.nis_no}", f"Moved from {old_office} to {new_office}", organization_id=organization_id)
        return jsonify(schemas.to_dict_staff(staff))

@app.get("/staff/<int:staff_id>/history")
def get_staff_history(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "main_admin"])
    if err: return err, code
    
    with next(get_db()) as db:
        stmt = select(models.PostingHistory).where(models.PostingHistory.staff_id == staff_id).order_by(models.PostingHistory.action_date.desc(), models.PostingHistory.created_at.desc())
        history = db.scalars(stmt).all()
        
        res = []
        for h in history:
            res.append({
                "id": h.id,
                "action_type": h.action_type,
                "from_office": h.from_office,
                "to_office": h.to_office,
                "action_date": h.action_date.isoformat() if h.action_date else None,
                "remarks": h.remarks,
                "created_at": h.created_at.isoformat() if h.created_at else None
            })
        return jsonify(res)

@app.get("/settings/staff-edit")
def get_staff_edit_settings():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    with next(get_db()) as db:
        # Filter by organization_id if present
        stmt = select(models.Staff)
        if organization_id:
            stmt = stmt.where(models.Staff.organization_id == organization_id)
            
        # We need to aggregate over the filtered set.
        # But wait, crud operations or simple scalars?
        # Let's construct queries manually since these are aggregates.
        
        q_max_rank = select(func.max(models.Staff.allow_edit_rank))
        q_max_dopp = select(func.max(models.Staff.allow_edit_dopp))
        q_max_login = select(func.max(models.Staff.allow_login))
        
        if organization_id:
            q_max_rank = q_max_rank.where(models.Staff.organization_id == organization_id)
            q_max_dopp = q_max_dopp.where(models.Staff.organization_id == organization_id)
            q_max_login = q_max_login.where(models.Staff.organization_id == organization_id)

        max_rank = db.scalar(q_max_rank) or 0
        max_dopp = db.scalar(q_max_dopp) or 0
        max_login = db.scalar(q_max_login)
        
        if max_login is None: max_login = 1 # Default to allowed if table empty?
        
        return jsonify({
            "allow_edit_rank": bool(max_rank),
            "allow_edit_dopp": bool(max_dopp),
            "allow_login": bool(max_login),
        })

@app.put("/settings/staff-edit")
def update_staff_edit_settings():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    data = request.get_json(force=True)
    allow_rank = bool(data.get("allow_edit_rank"))
    allow_dopp = bool(data.get("allow_edit_dopp"))
    allow_login = bool(data.get("allow_login"))
    
    with next(get_db()) as db:
        stmt = models.Staff.__table__.update().values(
            allow_edit_rank=1 if allow_rank else 0,
            allow_edit_dopp=1 if allow_dopp else 0,
            allow_login=1 if allow_login else 0,
        )
        if organization_id:
            stmt = stmt.where(models.Staff.organization_id == organization_id)
            
        db.execute(stmt)
        db.commit()
        crud.create_audit_log(
            db,
            "SETTINGS_UPDATE",
            "staff-edit",
            f"allow_edit_rank={allow_rank}, allow_edit_dopp={allow_dopp}, allow_login={allow_login}",
            organization_id=organization_id
        )
    return jsonify({"allow_edit_rank": allow_rank, "allow_edit_dopp": allow_dopp, "allow_login": allow_login})

@app.post("/change-password")
def change_password():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user_info = get_current_user()
    if not user_info: return jsonify({"detail": "Not authenticated"}), 401
    
    data = request.get_json(force=True)
    old_password = data.get("old_password")
    new_password = data.get("new_password")
    
    if not old_password or not new_password:
        return jsonify({"detail": "Old and new passwords are required"}), 400
        
    with next(get_db()) as db:
        # Identify if user is in User table or Staff table
        # We use ID and Sub (username/nis) to verify
        
        # 1. Try User table
        user_obj = db.query(models.User).filter(models.User.id == user_info["id"]).first()
        if user_obj and user_obj.username == user_info["sub"]:
             if not auth.verify_password(old_password, user_obj.password_hash):
                 return jsonify({"detail": "Incorrect old password"}), 400
             
             user_obj.password_hash = auth.get_password_hash(new_password)
             db.commit()
             crud.create_audit_log(db, "PASSWORD_CHANGE", user_obj.username, "Admin changed password", organization_id=user_obj.organization_id)
             return jsonify({"detail": "Password changed successfully"})

        # 2. Try Staff table
        staff = crud.get_staff(db, user_info["id"])
        if staff and staff.nis_no == user_info["sub"]:
            # Verify old password
            valid_old = False
            if staff.password_hash:
                 if auth.verify_password(old_password, staff.password_hash): valid_old = True
            elif old_password == staff.nis_no:
                 valid_old = True
                 
            if not valid_old:
                return jsonify({"detail": "Incorrect old password"}), 400
                
            staff.password_hash = auth.get_password_hash(new_password)
            db.commit()
            crud.create_audit_log(db, "PASSWORD_CHANGE", staff.nis_no, "User changed password", organization_id=staff.organization_id)
            return jsonify({"detail": "Password changed successfully"})

        return jsonify({"detail": "User record not found"}), 404

@app.get("/audit-logs")
def get_audit_logs():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "main_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    limit = request.args.get("limit", 100, type=int)
    offset = request.args.get("offset", 0, type=int)
    
    with next(get_db()) as db:
        logs = crud.list_audit_logs(db, limit=limit, offset=offset, organization_id=organization_id)
        return jsonify([schemas.to_dict_audit_log(l) for l in logs])

@app.get("/export/excel")
def export_excel():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user = get_current_user()
    if not user: return jsonify({"detail": "Not authenticated"}), 401
    
    organization_id = user.get("organization_id")
    
    try:
        with next(get_db()) as db:
            q = request.args.get("q")
            
            rank = [r for r in request.args.getlist("rank") if r.strip()]
            if not rank:
                rank = None
            
            office = [o for o in request.args.getlist("office") if o.strip()]
            if not office:
                office = None

            completeness = request.args.get("completeness")
            status = request.args.get("status", "active")
            dopp_order = request.args.get("dopp_order")
            columns_raw = request.args.get("columns")
            merge_name = request.args.get("merge_name") in ("1", "true", "True", "yes", "on")

            columns = []
            if columns_raw:
                columns = [c.strip() for c in columns_raw.split(",") if c and c.strip()]

            staff_list = crud.list_staff(
                db,
                q=q,
                rank=rank,
                office=office,
                completeness=completeness,
                status=status,
                dopp_order=dopp_order,
                limit=10000,
                offset=0,
                organization_id=organization_id
            )

            def tokenize_alpha_words(text: str) -> list[str]:
                if not text:
                    return []
                parts = []
                buf = []
                for ch in str(text).strip():
                    if ch.isalpha():
                        buf.append(ch)
                    else:
                        if buf:
                            parts.append("".join(buf))
                            buf = []
                if buf:
                    parts.append("".join(buf))
                return [p for p in parts if p]

            def initials_from_words(words: list[str]) -> str:
                return "".join([w[0].upper() for w in words if w])

            def normalize_rank_code(value: str) -> str:
                if not value:
                    return ""
                return "".join([ch for ch in str(value).upper() if ch.isalnum()])

            senior_ranks = {
                "DCG",
                "ACG",
                "CIS",
                "DCI",
                "ACI",
                "CSI",
                "SI",
                "DSI",
                "ASI1",
                "ASI2",
            }
            junior_ranks = {
                "II",
                "AII",
                "IA1",
                "IA2",
                "IA3",
            }

            rank_order = {
                "DCG": 1, "ACG": 2, "CIS": 3, "DCI": 4, "ACI": 5,
                "CSI": 6, "SI": 7, "DSI": 8, "ASI 1": 9, "ASI1": 9,
                "ASI 2": 10, "ASI2": 10, "II": 11, "AII": 12,
                "IA 1": 13, "IA1": 13, "IA 2": 14, "IA2": 14, "IA 3": 15, "IA3": 15
            }
            
            def get_rank_priority(staff):
                 r = normalize_rank_code(staff.rank)
                 if staff.rank and staff.rank.upper() in rank_order:
                     return rank_order[staff.rank.upper()]
                 for k, v in rank_order.items():
                     if normalize_rank_code(k) == r:
                         return v
                 return 999

            staff_list.sort(key=get_rank_priority)

            def merged_name_by_rank(staff) -> str:
                other_words = tokenize_alpha_words(staff.other_names or "")
                surname_full = (staff.surname or "").strip()
                surname_words = tokenize_alpha_words(staff.surname or "")

                other_initials = initials_from_words(other_words)
                surname_initials = initials_from_words(surname_words)

                rank_code = normalize_rank_code(staff.rank or "")

                if rank_code in senior_ranks:
                    if other_initials and surname_full:
                        return f"{other_initials} {surname_full}".strip()
                    return (surname_full or other_initials).strip()

                if rank_code in junior_ranks:
                    if not other_words:
                        return surname_full
                    first_name = other_words[0].upper()
                    rest_initials = initials_from_words(other_words[1:])
                    tail = f"{rest_initials}{surname_initials}".strip()
                    if tail:
                        return f"{first_name} {tail}".strip()
                    return first_name

                if surname_full:
                    return f"{surname_full}{(' ' + other_initials) if other_initials else ''}".strip()
                return other_initials

            rank_order = {
                "DCG": 1, "ACG": 2, "CIS": 3, "DCI": 4, "ACI": 5,
                "CSI": 6, "SI": 7, "DSI": 8, "ASI 1": 9, "ASI1": 9,
                "ASI 2": 10, "ASI2": 10, "II": 11, "AII": 12,
                "IA 1": 13, "IA1": 13, "IA 2": 14, "IA2": 14, "IA 3": 15, "IA3": 15
            }
            
            def get_rank_priority(staff):
                 r = normalize_rank_code(staff.rank)
                 if staff.rank and staff.rank.upper() in rank_order:
                     return rank_order[staff.rank.upper()]
                 for k, v in rank_order.items():
                     if normalize_rank_code(k) == r:
                         return v
                 return 999

            staff_list.sort(key=get_rank_priority)

            def get_value(staff, col_key: str):
                if col_key == "nis_no":
                    return staff.nis_no
                if col_key == "surname":
                    return staff.surname
                if col_key == "other_names":
                    return staff.other_names
                if col_key == "rank":
                    return staff.rank
                if col_key == "gender":
                    return staff.gender
                if col_key == "office":
                    return staff.office
                if col_key == "state":
                    return staff.state.name if staff.state else ""
                if col_key == "lga":
                    return staff.lga.name if staff.lga else ""
                if col_key == "phone_no":
                    return staff.phone_no
                if col_key == "qualification":
                    return staff.qualification
                if col_key == "dob":
                    return staff.dob.strftime('%d/%m/%Y') if staff.dob else ""
                if col_key == "dofa":
                    return staff.dofa.strftime('%d/%m/%Y') if staff.dofa else ""
                if col_key == "dopa":
                    return staff.dopa.strftime('%d/%m/%Y') if staff.dopa else ""
                if col_key == "dopp":
                    return staff.dopp.strftime('%d/%m/%Y') if staff.dopp else ""
                if col_key == "home_town":
                    return staff.home_town
                if col_key == "next_of_kin":
                    return staff.next_of_kin
                if col_key == "nok_phone":
                    return staff.nok_phone
                if col_key == "email":
                    return staff.email
                if col_key == "remark":
                    return staff.remark
                return ""

            label_map = {
                "nis_no": "NIS No",
                "surname": "Surname",
                "other_names": "Other Names",
                "rank": "Rank",
                "gender": "Gender",
                "office": "Office",
                "state": "State",
                "lga": "LGA",
                "phone_no": "Phone No",
                "qualification": "Qual",
                "dob": "DOB",
                "dofa": "DOFA",
                "dopa": "DOPA",
                "dopp": "DOPP",
                "home_town": "Home Town",
                "next_of_kin": "Next of Kin",
                "nok_phone": "NOK Phone",
                "email": "Email",
                "remark": "Remark",
            }

            if not columns:
                columns = ["nis_no", "surname", "other_names", "rank", "gender", "office", "state", "lga", "phone_no"]

            if "sn" in columns:
                columns.remove("sn")
            columns.insert(0, "sn")

            if merge_name and ("surname" in columns or "other_names" in columns):
                columns = [c for c in columns if c not in ("surname", "other_names")]
                name_col_key = "__name__"
                insert_at = 0
                if "nis_no" in columns:
                    insert_at = columns.index("nis_no") + 1
                columns.insert(insert_at, name_col_key)
                label_map[name_col_key] = "Name"
            
            label_map["sn"] = "S/N"
            
            main_title = "Visa/Residency Directorate"
            subtitle_text = ""
            office_title = None
            rank_title = None
            if office:
                if isinstance(office, list):
                    if len(office) == 1:
                        office_title = office[0]
                    else:
                        office_title = ", ".join(office)
                else:
                    office_title = office
            if rank:
                if isinstance(rank, list):
                    if len(rank) == 1:
                        rank_title = rank[0]
                    else:
                        rank_title = ", ".join(rank)
                else:
                    rank_title = rank
            if office_title:
                main_title = office_title
                subtitle_text = "Visa/Residency Directorate"
            elif rank_title:
                main_title = rank_title
                subtitle_text = "Visa/Residency Directorate"
            main_title = str(main_title)
            safe_filename = "".join([c for c in main_title if c.isalnum() or c in (' ', '-', '_')]).strip().replace(' ', '_')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Staff List"
            
            font_style = Font(name='Liberation Sans', size=10)
            header_font = Font(name='Liberation Sans', size=12, bold=True)
            title_font = Font(name='Liberation Sans', size=14, bold=True)
            subtitle_font = Font(name='Liberation Sans', size=10, italic=False)
            center_align = Alignment(horizontal='center', vertical='center')
            
            current_row = 1
            ws.cell(row=current_row, column=1, value=main_title)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
            cell = ws.cell(row=current_row, column=1)
            cell.font = title_font
            cell.alignment = center_align
            current_row += 1
            
            if subtitle_text:
                ws.cell(row=current_row, column=1, value=subtitle_text)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
                cell = ws.cell(row=current_row, column=1)
                cell.font = subtitle_font
                cell.alignment = center_align
                current_row += 1
                
            headers = [label_map.get(c, c).upper() for c in columns]
            ws.append(headers)
            header_row_idx = current_row
            for cell in ws[header_row_idx]:
                cell.font = header_font
            for idx, staff in enumerate(staff_list, start=1):
                row = []
                for col_key in columns:
                    if col_key == "sn":
                        row.append(idx)
                    elif col_key == "__name__":
                        row.append(merged_name_by_rank(staff))
                    else:
                        row.append(get_value(staff, col_key))
                ws.append(row)
                row_idx = idx + header_row_idx
                if row_idx % 2 == 0:
                    fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    for cell in ws[row_idx]:
                        cell.fill = fill
                for cell in ws[row_idx]:
                    cell.font = font_style
            ws.append([])
            footer_cell = ws.cell(row=ws.max_row + 1, column=1, value=f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            footer_cell.font = Font(name='Liberation Sans', size=8, italic=True)
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"{safe_filename}_{datetime.now().strftime('%Y%m%d')}.xlsx")
    except Exception as e:
        print("Export Excel error:", e)
        return jsonify({"detail": f"Export Excel failed: {str(e)}"}), 500

@app.get("/export/pdf")
def export_pdf():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user = get_current_user()
    if not user: return jsonify({"detail": "Not authenticated"}), 401
    
    organization_id = user.get("organization_id")
    
    try:
        with next(get_db()) as db:
            q = request.args.get("q")
            
            rank = [r for r in request.args.getlist("rank") if r.strip()]
            if not rank:
                rank = None
            
            office = [o for o in request.args.getlist("office") if o.strip()]
            if not office:
                office = None

            completeness = request.args.get("completeness")
            status = request.args.get("status", "active")
            dopp_order = request.args.get("dopp_order")
            columns_raw = request.args.get("columns")
            merge_name = request.args.get("merge_name") in ("1", "true", "True", "yes", "on")

            columns = []
            if columns_raw:
                columns = [c.strip() for c in columns_raw.split(",") if c and c.strip()]
            if not columns:
                columns = ["nis_no", "surname", "other_names", "rank", "gender", "office", "state", "lga", "phone_no"]

            if "sn" in columns:
                columns.remove("sn")
            columns.insert(0, "sn")

            staff_list = crud.list_staff(
                db,
                q=q,
                rank=rank,
                office=office,
                completeness=completeness,
                status=status,
                dopp_order=dopp_order,
                limit=10000,
                offset=0,
                organization_id=organization_id
            )

            def tokenize_alpha_words(text: str) -> list[str]:
                if not text:
                    return []
                parts = []
                buf = []
                for ch in str(text).strip():
                    if ch.isalpha():
                        buf.append(ch)
                    else:
                        if buf:
                            parts.append("".join(buf))
                            buf = []
                if buf:
                    parts.append("".join(buf))
                return [p for p in parts if p]

            def initials_from_words(words: list[str]) -> str:
                return "".join([w[0].upper() for w in words if w])

            def normalize_rank_code(value: str) -> str:
                if not value:
                    return ""
                return "".join([ch for ch in str(value).upper() if ch.isalnum()])

            senior_ranks = {"DCG","ACG","CIS","DCI","ACI","CSI","SI","DSI","ASI1","ASI2"}
            junior_ranks = {"II","AII","IA1","IA2","IA3"}

            def merged_name_by_rank(staff) -> str:
                other_words = tokenize_alpha_words(staff.other_names or "")
                surname_full = (staff.surname or "").strip()
                surname_words = tokenize_alpha_words(staff.surname or "")
                other_initials = initials_from_words(other_words)
                surname_initials = initials_from_words(surname_words)
                rank_code = normalize_rank_code(staff.rank or "")
                if rank_code in senior_ranks:
                    if other_initials and surname_full:
                        return f"{other_initials} {surname_full}".strip()
                    return (surname_full or other_initials).strip()
                if rank_code in junior_ranks:
                    if not other_words:
                        return surname_full
                    first_name = other_words[0].upper()
                    rest_initials = initials_from_words(other_words[1:])
                    tail = f"{rest_initials}{surname_initials}".strip()
                    if tail:
                        return f"{first_name} {tail}".strip()
                    return first_name
                if surname_full:
                    return f"{surname_full}{(' ' + other_initials) if other_initials else ''}".strip()
                return other_initials

            def get_value(staff, col_key: str):
                if col_key == "nis_no":
                    return staff.nis_no or ""
                if col_key == "surname":
                    return staff.surname or ""
                if col_key == "other_names":
                    return staff.other_names or ""
                if col_key == "rank":
                    return staff.rank or ""
                if col_key == "gender":
                    return staff.gender or ""
                if col_key == "office":
                    return staff.office or ""
                if col_key == "state":
                    return staff.state.name if staff.state else ""
                if col_key == "lga":
                    return staff.lga.name if staff.lga else ""
                if col_key == "phone_no":
                    return staff.phone_no or ""
                if col_key == "qualification":
                    return staff.qualification or ""
                if col_key == "dob":
                    return staff.dob.strftime('%d/%m/%Y') if staff.dob else ""
                if col_key == "dofa":
                    return staff.dofa.strftime('%d/%m/%Y') if staff.dofa else ""
                if col_key == "dopa":
                    return staff.dopa.strftime('%d/%m/%Y') if staff.dopa else ""
                if col_key == "dopp":
                    return staff.dopp.strftime('%d/%m/%Y') if staff.dopp else ""
                if col_key == "home_town":
                    return staff.home_town or ""
                if col_key == "next_of_kin":
                    return staff.next_of_kin or ""
                if col_key == "nok_phone":
                    return staff.nok_phone or ""
                if col_key == "email":
                    return staff.email or ""
                if col_key == "remark":
                    return staff.remark or ""
                return ""

            label_map = {
                "nis_no": "NIS No",
                "surname": "Surname",
                "other_names": "Other Names",
                "rank": "Rank",
                "gender": "Gender",
                "office": "Office",
                "state": "State",
                "lga": "LGA",
                "phone_no": "Phone No",
                "qualification": "Qual",
                "dob": "DOB",
                "dofa": "DOFA",
                "dopa": "DOPA",
                "dopp": "DOPP",
                "home_town": "Home Town",
                "next_of_kin": "Next of Kin",
                "nok_phone": "NOK Phone",
                "email": "Email",
                "remark": "Remark",
                "sn": "S/N",
            }

            headers_keys = list(columns)
            if merge_name and ("surname" in headers_keys or "other_names" in headers_keys):
                headers_keys = [c for c in headers_keys if c not in ("surname", "other_names")]
                name_col_key = "__name__"
                insert_at = 0
                if "nis_no" in headers_keys:
                    insert_at = headers_keys.index("nis_no") + 1
                headers_keys.insert(insert_at, name_col_key)
                label_map[name_col_key] = "Name"

            data_table = [[label_map.get(k, k).upper() for k in headers_keys]]
            for idx, staff in enumerate(staff_list, start=1):
                row = []
                for k in headers_keys:
                    if k == "sn":
                        row.append(str(idx))
                    elif k == "__name__":
                        row.append(merged_name_by_rank(staff))
                    else:
                        row.append(get_value(staff, k))
                data_table.append(row)

            out = io.BytesIO()
            # Use A3 Landscape for more width
            doc = SimpleDocTemplate(out, pagesize=landscape(A3), topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
            styles = getSampleStyleSheet()
            elements = []
            
            main_title = "Visa/Residency Directorate"
            subtitle_text = ""
            office_title = None
            rank_title = None
            if office:
                if isinstance(office, list):
                    if len(office) == 1:
                        office_title = office[0]
                    else:
                        office_title = ", ".join(office)
                else:
                    office_title = office
            if rank:
                if isinstance(rank, list):
                    if len(rank) == 1:
                        rank_title = rank[0]
                    else:
                        rank_title = ", ".join(rank)
                else:
                    rank_title = rank
            if office_title:
                main_title = office_title
                subtitle_text = "Visa/Residency Directorate"
            elif rank_title:
                main_title = rank_title
                subtitle_text = "Visa/Residency Directorate"
            main_title = str(main_title)
            title_style = styles["Title"]
            title_style.fontSize = 14
            elements.append(Paragraph(main_title, title_style))
            
            if subtitle_text:
                 subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=1, fontSize=10)
                 elements.append(Paragraph(subtitle_text, subtitle_style))
            
            safe_filename = "".join([c for c in main_title if c.isalnum() or c in (' ', '-', '_')]).strip().replace(' ', '_')

            elements.append(Spacer(1, 0.2 * inch))
            
            # A3 Landscape width is approx 1190 points. With 30 margins, we have ~1130 available.
            avail_width = 1130

            font_size = 7
            if len(headers_keys) > 12:
                font_size = 6

            char_width = font_size * 0.6
            min_width_map = {
                "sn": 25,
                "rank": 35,
                "gender": 40,
                "nis_no": 45,
                "qualification": 40,
                "dob": 55, "dofa": 55, "dopa": 55, "dopp": 55,
                "phone_no": 65,
                "state": 60, "lga": 60,
                "grade_level": 30, "step": 25,
                "nok_phone": 65,
            }
            max_len_map = {}
            for k in headers_keys:
                header_text = label_map.get(k, k)
                max_len_map[k] = len(str(header_text))
            if "sn" in headers_keys:
                max_len_map["sn"] = max(max_len_map.get("sn", 0), len(str(len(staff_list))))
            for staff in staff_list:
                for k in headers_keys:
                    if k == "sn":
                        s = ""
                    elif k == "__name__":
                        s = merged_name_by_rank(staff)
                    else:
                        s = str(get_value(staff, k))
                    max_len_map[k] = max(max_len_map.get(k, 0), len(s or ""))
            raw_widths = []
            for k in headers_keys:
                base = (max_len_map.get(k, 1) + 2) * char_width
                min_w = min_width_map.get(k, 40)
                raw_widths.append(max(base, min_w))
            total_raw = sum(raw_widths)
            if total_raw > 0:
                scale = avail_width / total_raw
            else:
                scale = 1.0
            final_widths = [w * scale for w in raw_widths]

            font_size = 7
            if len(headers_keys) > 12:
                font_size = 6

            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=font_size,
                leading=font_size + 1,
                alignment=0
            )
            
            formatted_data = [data_table[0]]
            
            for idx, staff in enumerate(staff_list, start=1):
                row = []
                for k in headers_keys:
                    val = ""
                    if k == "sn":
                        val = str(idx)
                    elif k == "__name__":
                        val = merged_name_by_rank(staff)
                    else:
                        val = str(get_value(staff, k))
                    if val:
                        row.append(Paragraph(val, cell_style))
                    else:
                        row.append("")
                formatted_data.append(row)

            table = Table(formatted_data, repeatRows=1, colWidths=final_widths)
            
            style = TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), font_size + 1),
                ("FONTSIZE", (0, 1), (-1, -1), font_size),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ])
            table.setStyle(style)
            for i in range(1, len(formatted_data)):
                if i % 2 == 0:
                    table.setStyle(TableStyle([("BACKGROUND", (0, i), (-1, i), colors.whitesmoke)]))
            elements.append(table)
            
            def footer(canvas, doc):
                canvas.saveState()
                canvas.setFont('Helvetica-Oblique', 8)
                page_num = canvas.getPageNumber()
                text = f"Page {page_num} | Generated on {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                canvas.drawRightString(landscape(letter)[0] - 30, 20, text)
                canvas.restoreState()

            doc.build(elements, onFirstPage=footer, onLaterPages=footer)
            out.seek(0)
            return send_file(out, mimetype="application/pdf", as_attachment=True, download_name=f"{safe_filename}_{datetime.now().strftime('%Y%m%d')}.pdf")
    except Exception as e:
        print("Export PDF error:", e)
        return jsonify({"detail": f"Export PDF failed: {str(e)}"}), 500

@app.post("/staff/<int:staff_id>/exit-request")
def request_exit(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user = get_current_user()
    if not user: return jsonify({"detail": "Not authenticated"}), 401
    
    data = request.get_json(force=True)
    exit_date = parse_date_value(data.get("exit_date"))
    exit_mode = data.get("exit_mode")
    
    if not exit_date or not exit_mode:
        return jsonify({"detail": "Exit date and mode required"}), 400
        
    with next(get_db()) as db:
        staff = crud.get_staff(db, staff_id)
        if not staff: return jsonify({"detail": "Not found"}), 404
        
        # Check permission (office admin only for own office)
        if user["role"] == "office_admin":
            admin_staff = crud.get_staff(db, user["id"])
            if not admin_staff or admin_staff.office != staff.office:
                return jsonify({"detail": "Permission denied"}), 403
        
        staff.out_request_status = "Pending"
        staff.out_request_date = exit_date
        staff.out_request_reason = exit_mode
        db.commit()
        crud.create_audit_log(db, "EXIT_REQUEST", f"Staff: {staff.nis_no}", f"Requested exit: {exit_mode} on {exit_date}")
        return jsonify({"detail": "Request submitted"})

@app.post("/staff/<int:staff_id>/exit-approve")
def approve_exit(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "main_admin"])
    if err: return err, code
    
    with next(get_db()) as db:
        staff = crud.get_staff(db, staff_id)
        if not staff: return jsonify({"detail": "Not found"}), 404
        
        if not staff.out_request_status:
            return jsonify({"detail": "No pending request"}), 400
            
        staff.exit_date = staff.out_request_date
        staff.exit_mode = staff.out_request_reason
        staff.out_request_status = None
        
        # Log History
        history = models.PostingHistory(
            staff_id=staff.id,
            action_type="EXIT",
            from_office=staff.office,
            to_office=None,
            action_date=staff.exit_date,
            remarks=staff.exit_mode
        )
        db.add(history)

        staff.out_request_date = None
        staff.out_request_reason = None
        
        db.commit()
        crud.create_audit_log(db, "EXIT_APPROVE", f"Staff: {staff.nis_no}", "Approved exit request")
        return jsonify(schemas.to_dict_staff(staff))

@app.post("/staff/<int:staff_id>/exit-reject")
def reject_exit(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "main_admin"])
    if err: return err, code
    
    with next(get_db()) as db:
        staff = crud.get_staff(db, staff_id)
        if not staff: return jsonify({"detail": "Not found"}), 404
        
        staff.out_request_status = None
        staff.out_request_date = None
        staff.out_request_reason = None
        
        db.commit()
        crud.create_audit_log(db, "EXIT_REJECT", f"Staff: {staff.nis_no}", "Rejected exit request")
        return jsonify({"detail": "Request rejected"})

@app.post("/staff/<int:staff_id>/undo-exit")
def undo_exit(staff_id: int):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "main_admin"])
    if err: return err, code
    
    with next(get_db()) as db:
        staff = crud.get_staff(db, staff_id)
        if not staff: return jsonify({"detail": "Not found"}), 404
        
        if not staff.exit_date:
            return jsonify({"detail": "Staff is not exited"}), 400
            
        # Log History
        history = models.PostingHistory(
            staff_id=staff.id,
            action_type="UNDO_EXIT",
            from_office=None,
            to_office=staff.office,
            action_date=date.today(),
            remarks="Undo Exit"
        )
        db.add(history)

        staff.exit_date = None
        staff.exit_mode = None
        # Also clear request fields if they linger
        staff.out_request_status = None
        staff.out_request_date = None
        staff.out_request_reason = None
        
        db.commit()
        crud.create_audit_log(db, "UNDO_EXIT", f"Staff: {staff.nis_no}", "Undid exit/posting out")
        return jsonify(schemas.to_dict_staff(staff))

@app.get("/admin/edit-requests")
def list_edit_requests():
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "main_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    with next(get_db()) as db:
        stmt = select(models.StaffEditRequest).join(models.Staff).where(models.StaffEditRequest.status == "pending")
        if organization_id:
            stmt = stmt.where(models.Staff.organization_id == organization_id)
        stmt = stmt.order_by(models.StaffEditRequest.created_at.desc())
        
        reqs = db.scalars(stmt).all()
        
        res = []
        for r in reqs:
            res.append({
                "id": r.id,
                "staff_id": r.staff_id,
                "staff_name": f"{r.staff.surname} {r.staff.other_names}",
                "staff_nis": r.staff.nis_no,
                "data": json.loads(r.data),
                "created_at": r.created_at.isoformat(),
            })
        return jsonify(res)

@app.post("/admin/edit-requests/<int:req_id>/approve")
def approve_edit_request(req_id):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "main_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    with next(get_db()) as db:
        try:
            req = db.get(models.StaffEditRequest, req_id)
            if not req:
                return jsonify({"detail": "Not found"}), 404
            if req.status != "pending":
                return jsonify({"detail": "Request not pending"}), 400
            
            staff = db.get(models.Staff, req.staff_id)
            if not staff:
                return jsonify({"detail": "Staff not found"}), 404
                
            if organization_id and staff.organization_id != organization_id:
                return jsonify({"detail": "Permission denied: Different Organization"}), 403
            
            data = json.loads(req.data)
            for k in ("dofa", "dopa", "dopp", "dob", "exit_date"):
                if k in data and data[k]:
                    parsed = parse_date_value(data[k])
                    if parsed is None and data[k] not in (None, "", 0):
                        return jsonify({"detail": f"Invalid date for {k}"}), 400
                    data[k] = parsed
            
            crud.update_staff(db, staff, data)
            
            req.status = "approved"
            req.reviewed_by = user.get("sub")
            req.reviewed_at = func.now()
            
            crud.create_audit_log(db, "APPROVE_EDIT", f"Staff: {staff.nis_no}", f"Approved edit request {req_id}", organization_id=organization_id)
            db.commit()
            return jsonify({"detail": "Request approved and applied"})
        except Exception as e:
            import traceback
            db.rollback()
            print("Approve edit request error:", e)
            print(traceback.format_exc())
            return jsonify({"detail": f"Approve edit failed: {str(e)}"}), 500

@app.post("/admin/edit-requests/<int:req_id>/reject")
def reject_edit_request(req_id):
    if STARTUP_ERROR: return jsonify({"detail": STARTUP_ERROR}), 500
    user, err, code = require_role(["super_admin", "main_admin"])
    if err: return err, code
    
    organization_id = user.get("organization_id")
    
    with next(get_db()) as db:
        req = db.get(models.StaffEditRequest, req_id)
        if not req: return jsonify({"detail": "Not found"}), 404
        if req.status != "pending": return jsonify({"detail": "Request not pending"}), 400
        
        staff = db.get(models.Staff, req.staff_id)
        if staff and organization_id and staff.organization_id != organization_id:
             return jsonify({"detail": "Permission denied: Different Organization"}), 403
        
        req.status = "rejected"
        req.reviewed_by = user.get("sub")
        req.reviewed_at = func.now()
        
        crud.create_audit_log(db, "REJECT_EDIT", f"Request {req_id}", "Rejected edit request", organization_id=organization_id)
        db.commit()
        return jsonify({"detail": "Request rejected"})
